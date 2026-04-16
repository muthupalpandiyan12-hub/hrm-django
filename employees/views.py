from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from .models import Employee, Department, SalaryHistory
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


@login_required
def employee_list(request):
    employees = Employee.objects.select_related('department').all()
    return render(request, 'employees/list.html', {'employees': employees})


@login_required
def employee_add(request):
    departments = Department.objects.all()
    if request.method == 'POST':
        try:
            Employee.objects.create(
                employee_id=request.POST['employee_id'],
                name=request.POST['name'],
                email=request.POST['email'],
                phone=request.POST.get('phone', ''),
                gender=request.POST.get('gender', ''),
                date_of_birth=request.POST.get('date_of_birth') or None,
                address=request.POST.get('address', ''),
                department_id=request.POST.get('department') or None,
                position=request.POST.get('position', ''),
                salary=request.POST.get('salary', 0),
                status=request.POST.get('status', 'active'),
            )
            messages.success(request, 'Employee added successfully.')
            return redirect('employee_list')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    return render(request, 'employees/form.html', {'departments': departments, 'action': 'Add'})


@login_required
def employee_edit(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    departments = Department.objects.all()
    if request.method == 'POST':
        try:
            old_salary = employee.salary
            employee.employee_id = request.POST['employee_id']
            employee.name = request.POST['name']
            employee.email = request.POST['email']
            employee.phone = request.POST.get('phone', '')
            employee.gender = request.POST.get('gender', '')
            employee.date_of_birth = request.POST.get('date_of_birth') or None
            employee.address = request.POST.get('address', '')
            employee.department_id = request.POST.get('department') or None
            employee.position = request.POST.get('position', '')
            new_salary = request.POST.get('salary', 0)
            employee.salary = new_salary
            employee.status = request.POST.get('status', 'active')
            employee.save()
            # Record salary change if salary changed
            import decimal
            try:
                new_salary_dec = decimal.Decimal(str(new_salary))
            except Exception:
                new_salary_dec = old_salary
            if new_salary_dec != old_salary:
                import datetime
                SalaryHistory.objects.create(
                    employee=employee,
                    old_salary=old_salary,
                    new_salary=new_salary_dec,
                    effective_date=datetime.date.today(),
                    reason=request.POST.get('salary_change_reason', ''),
                    reason_type=request.POST.get('salary_change_reason_type', 'other'),
                )
            messages.success(request, 'Employee updated successfully.')
            return redirect('employee_list')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    return render(request, 'employees/form.html', {'employee': employee, 'departments': departments, 'action': 'Edit'})


@login_required
def salary_history(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    history = SalaryHistory.objects.filter(employee=employee)
    return render(request, 'employees/salary_history.html', {'employee': employee, 'history': history})


@login_required
def employee_delete(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.delete()
        messages.success(request, 'Employee deleted successfully.')
        return redirect('employee_list')
    return render(request, 'employees/confirm_delete.html', {'employee': employee})


@login_required
def employee_detail(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    return render(request, 'employees/detail.html', {'employee': employee})


@login_required
def employee_export(request):
    employees = Employee.objects.select_related('department').all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    header_fill  = PatternFill("solid", fgColor="E60000")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    center       = Alignment(horizontal="center", vertical="center")
    border       = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'),  bottom=Side(style='thin'))

    headers = ['#','Employee ID','Name','Email','Phone','Gender',
               'Department','Position','Salary','Date Joined','Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = center; cell.border = border

    alt_fill = PatternFill("solid", fgColor="f1f5f9")
    for i, emp in enumerate(employees, 1):
        ws.append([
            i, emp.employee_id, emp.name, emp.email,
            emp.phone or '—', emp.get_gender_display() or '—',
            str(emp.department) if emp.department else '—',
            emp.position or '—', float(emp.salary),
            str(emp.date_joined), emp.status.upper()
        ])
        for cell in ws[i+1]:
            cell.border = border; cell.alignment = Alignment(vertical="center")
            if i % 2 == 0: cell.fill = alt_fill

    for col, w in zip(['A','B','C','D','E','F','G','H','I','J','K'],
                      [5,14,22,24,14,10,16,18,12,14,12]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Employees.xlsx"'
    wb.save(response)
    return response
