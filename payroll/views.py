from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from employees.models import Employee
from .models import Payroll
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT


MONTHS = {1:'January',2:'February',3:'March',4:'April',5:'May',6:'June',
          7:'July',8:'August',9:'September',10:'October',11:'November',12:'December'}


@login_required
def payroll_list(request):
    month = int(request.GET.get('month', datetime.date.today().month))
    year  = int(request.GET.get('year',  datetime.date.today().year))
    payrolls = Payroll.objects.filter(month=month, year=year).select_related('employee')
    return render(request, 'payroll/list.html', {
        'payrolls': payrolls, 'month': month, 'year': year,
        'months': range(1, 13),
        'years': range(datetime.date.today().year - 2, datetime.date.today().year + 1),
    })


@login_required
def payroll_generate(request):
    employees = Employee.objects.filter(status='active')
    if request.method == 'POST':
        employee_id = request.POST['employee']
        month = request.POST['month']
        year  = request.POST['year']
        employee = get_object_or_404(Employee, pk=employee_id)
        try:
            basic_salary = float(request.POST.get('basic_salary') or employee.salary or 0)
            allowances   = float(request.POST.get('allowances') or 0)
            deductions   = float(request.POST.get('deductions') or 0)
            payroll, created = Payroll.objects.update_or_create(
                employee=employee, month=int(month), year=int(year),
                defaults={
                    'basic_salary': basic_salary,
                    'allowances':   allowances,
                    'deductions':   deductions,
                    'status': request.POST.get('status', 'draft'),
                }
            )
            messages.success(request, f'Payroll {"created" if created else "updated"} successfully.')
            return redirect('payroll_list')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    return render(request, 'payroll/generate.html', {
        'employees': employees,
        'months': range(1, 13),
        'years': range(datetime.date.today().year - 2, datetime.date.today().year + 1),
        'today': datetime.date.today(),
    })


@login_required
def payroll_delete(request, pk):
    payroll = get_object_or_404(Payroll, pk=pk)
    if request.method == 'POST':
        payroll.delete()
        messages.success(request, 'Payroll record deleted.')
        return redirect('payroll_list')
    return render(request, 'payroll/confirm_delete.html', {'payroll': payroll})


# ─── PDF PAYSLIP ─────────────────────────────────────────────────────────────

# Vi brand colors
VI_RED   = colors.HexColor('#E60000')
VI_BLACK = colors.HexColor('#1A1A1A')
VI_GREY  = colors.HexColor('#F5F5F5')
VI_LIGHT = colors.HexColor('#FFF0F0')
VI_LINE  = colors.HexColor('#DDDDDD')
TXT_DARK = colors.HexColor('#1A1A1A')
TXT_GREY = colors.HexColor('#666666')


@login_required
def payroll_pdf(request, pk):
    payroll  = get_object_or_404(Payroll, pk=pk)
    employee = payroll.employee

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="Payslip_{employee.employee_id}_{MONTHS[payroll.month]}_{payroll.year}.pdf"'
    )

    doc = SimpleDocTemplate(
        response, pagesize=A4,
        topMargin=2*cm, bottomMargin=2*cm,
        leftMargin=2.2*cm, rightMargin=2.2*cm
    )
    story = []

    # ── Company Header ──────────────────────────────────────
    story.append(Spacer(1, 0.3*cm))

    company_style = ParagraphStyle(
        'company', fontSize=24, fontName='Helvetica-Bold',
        textColor=VI_RED, alignment=TA_CENTER,
        leading=32, spaceAfter=0, spaceBefore=0
    )
    tagline_style = ParagraphStyle(
        'tagline', fontSize=10, fontName='Helvetica',
        textColor=TXT_GREY, alignment=TA_CENTER,
        leading=16, spaceAfter=0, spaceBefore=0
    )

    story.append(Paragraph("HRM System", company_style))
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph("Human Resource Management", tagline_style))
    story.append(Spacer(1, 0.4*cm))

    # Red divider line
    story.append(HRFlowable(width="100%", thickness=3, color=VI_RED, spaceAfter=6))
    story.append(Spacer(1, 0.2*cm))

    # ── Payslip Title ──────────────────────────────────────
    title_style = ParagraphStyle(
        'title', fontSize=13, fontName='Helvetica-Bold',
        textColor=TXT_DARK, alignment=TA_CENTER,
        leading=20, spaceAfter=0, spaceBefore=0
    )
    period_style = ParagraphStyle(
        'period', fontSize=10, fontName='Helvetica',
        textColor=TXT_GREY, alignment=TA_CENTER,
        leading=16, spaceAfter=0, spaceBefore=0
    )
    story.append(Paragraph("PAYSLIP", title_style))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(f"{MONTHS[payroll.month].upper()} {payroll.year}", period_style))
    story.append(Spacer(1, 0.6*cm))

    # ── Employee Details Table ─────────────────────────────
    sec_header = ParagraphStyle(
        'secH', fontSize=10, fontName='Helvetica-Bold',
        textColor=VI_RED
    )
    field_style = ParagraphStyle('fld', fontSize=9, fontName='Helvetica-Bold', textColor=TXT_GREY)
    val_style   = ParagraphStyle('val', fontSize=9, fontName='Helvetica',      textColor=TXT_DARK)

    emp_data = [
        [Paragraph('EMPLOYEE DETAILS', sec_header), ''],
        [Paragraph('Employee Name',  field_style), Paragraph(employee.name or '—',                              val_style)],
        [Paragraph('Employee ID',    field_style), Paragraph(employee.employee_id or '—',                       val_style)],
        [Paragraph('Department',     field_style), Paragraph(str(employee.department) if employee.department else '—', val_style)],
        [Paragraph('Position',       field_style), Paragraph(employee.position or '—',                          val_style)],
        [Paragraph('Email',          field_style), Paragraph(employee.email or '—',                             val_style)],
        [Paragraph('Pay Period',     field_style), Paragraph(f"{MONTHS[payroll.month]} {payroll.year}",         val_style)],
        [Paragraph('Payment Status', field_style), Paragraph(payroll.status.upper(),                            val_style)],
    ]

    emp_table = Table(emp_data, colWidths=[5*cm, 11.5*cm])
    emp_table.setStyle(TableStyle([
        # Section header row
        ('SPAN',       (0, 0), (1, 0)),
        ('BACKGROUND', (0, 0), (1, 0), VI_LIGHT),
        ('TOPPADDING', (0, 0), (1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (1, 0), 10),
        ('LEFTPADDING',   (0, 0), (1, 0), 12),
        # Data rows
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, VI_GREY]),
        ('TOPPADDING',    (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('LEFTPADDING',   (0, 1), (-1, -1), 12),
        ('RIGHTPADDING',  (0, 1), (-1, -1), 12),
        # Borders
        ('BOX',      (0, 0), (-1, -1), 1, VI_LINE),
        ('LINEBELOW',(0, 0), (-1, -2), 0.5, VI_LINE),
        # Left accent bar on section header
        ('LINEAFTER', (0, 0), (0, -1), 0.5, VI_LINE),
    ]))
    story.append(emp_table)
    story.append(Spacer(1, 0.6*cm))

    # ── Salary Breakdown Table ─────────────────────────────
    col_hdr = ParagraphStyle('ch', fontSize=9, fontName='Helvetica-Bold', textColor=colors.white)

    salary_data = [
        [Paragraph('SALARY BREAKDOWN', sec_header), '', ''],
        [Paragraph('Description', col_hdr), Paragraph('Type', col_hdr), Paragraph('Amount (Rs.)', col_hdr)],
        ['Basic Salary', 'Earning',   f"Rs. {payroll.basic_salary:,.2f}"],
        ['Allowances',   'Earning',   f"Rs. {payroll.allowances:,.2f}"],
        ['Deductions',   'Deduction', f"Rs. {payroll.deductions:,.2f}"],
        ['NET SALARY',   '',          f"Rs. {payroll.net_salary:,.2f}"],
    ]

    col_w = [7*cm, 4.25*cm, 5.25*cm]
    salary_table = Table(salary_data, colWidths=col_w)
    salary_table.setStyle(TableStyle([
        # Section header
        ('SPAN',       (0, 0), (2, 0)),
        ('BACKGROUND', (0, 0), (2, 0), VI_LIGHT),
        ('TOPPADDING', (0, 0), (2, 0), 10),
        ('BOTTOMPADDING', (0, 0), (2, 0), 10),
        ('LEFTPADDING',   (0, 0), (2, 0), 12),

        # Column header row (row 1)
        ('BACKGROUND',    (0, 1), (2, 1), VI_BLACK),
        ('TOPPADDING',    (0, 1), (2, 1), 9),
        ('BOTTOMPADDING', (0, 1), (2, 1), 9),
        ('LEFTPADDING',   (0, 1), (2, 1), 12),
        ('RIGHTPADDING',  (0, 1), (2, 1), 12),

        # Data rows
        ('FONTNAME',  (0, 2), (-1, -2), 'Helvetica'),
        ('FONTSIZE',  (0, 0), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 2), (-1, -2), [colors.white, VI_GREY]),
        ('TOPPADDING',    (0, 2), (-1, -2), 9),
        ('BOTTOMPADDING', (0, 2), (-1, -2), 9),
        ('LEFTPADDING',   (0, 2), (-1, -2), 12),
        ('RIGHTPADDING',  (0, 2), (-1, -2), 12),

        # NET SALARY row (last row)
        ('BACKGROUND', (0, -1), (-1, -1), VI_BLACK),
        ('TEXTCOLOR',  (0, -1), (-1, -1), colors.white),
        ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING',    (0, -1), (-1, -1), 11),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 11),
        ('LEFTPADDING',   (0, -1), (-1, -1), 12),
        ('RIGHTPADDING',  (0, -1), (-1, -1), 12),

        # Alignment
        ('ALIGN',  (2, 1), (2, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # Borders
        ('BOX',      (0, 0), (-1, -1), 1, VI_LINE),
        ('LINEBELOW',(0, 0), (-1, -2), 0.5, VI_LINE),
        ('LINEBEFORE', (1, 1), (1, -1), 0.5, VI_LINE),
        ('LINEBEFORE', (2, 1), (2, -1), 0.5, VI_LINE),
    ]))
    story.append(salary_table)
    story.append(Spacer(1, 1.2*cm))

    # ── Footer ──────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=1, color=VI_LINE))
    story.append(Spacer(1, 0.3*cm))
    footer_style = ParagraphStyle(
        'footer', fontSize=8, fontName='Helvetica',
        textColor=TXT_GREY, alignment=TA_CENTER, leading=13
    )
    story.append(Paragraph(
        "This is a system-generated payslip and does not require a signature.", footer_style
    ))
    story.append(Paragraph(
        f"Generated on {datetime.date.today().strftime('%d %B %Y')}  |  HRM System", footer_style
    ))

    doc.build(story)
    return response


# ─── EXPORT PAYROLL TO EXCEL ─────────────────────────────────────────────────

@login_required
def payroll_export(request):
    month = int(request.GET.get('month', datetime.date.today().month))
    year  = int(request.GET.get('year',  datetime.date.today().year))
    payrolls = Payroll.objects.filter(month=month, year=year).select_related('employee')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Payroll {MONTHS[month]} {year}"

    # Header row style
    header_fill   = PatternFill("solid", fgColor="1e40af")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    header_align  = Alignment(horizontal="center", vertical="center")
    border        = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    headers = ['#', 'Employee ID', 'Name', 'Department', 'Position',
               'Basic Salary', 'Allowances', 'Deductions', 'Net Salary', 'Status']
    ws.append(headers)

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align
        cell.border    = border

    # Data rows
    alt_fill = PatternFill("solid", fgColor="f1f5f9")
    for i, p in enumerate(payrolls, start=1):
        row = [
            i, p.employee.employee_id, p.employee.name,
            str(p.employee.department) if p.employee.department else '—',
            p.employee.position or '—',
            float(p.basic_salary), float(p.allowances),
            float(p.deductions),   float(p.net_salary),
            p.status.upper()
        ]
        ws.append(row)
        if i % 2 == 0:
            for cell in ws[i + 1]:
                cell.fill = alt_fill
        for cell in ws[i + 1]:
            cell.border    = border
            cell.alignment = Alignment(vertical="center")

    # Column widths
    for col, width in zip(['A','B','C','D','E','F','G','H','I','J'],
                          [5, 14, 22, 16, 18, 14, 14, 14, 14, 12]):
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Payroll_{MONTHS[month]}_{year}.xlsx"'
    wb.save(response)
    return response
