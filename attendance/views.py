from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from employees.models import Employee
from .models import Attendance
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


@login_required
def attendance_list(request):
    date_filter = request.GET.get('date', str(datetime.date.today()))
    attendances = Attendance.objects.filter(date=date_filter).select_related('employee')
    employees = Employee.objects.filter(status='active')
    return render(request, 'attendance/list.html', {
        'attendances': attendances,
        'date_filter': date_filter,
        'employees': employees,
    })


@login_required
def attendance_mark(request):
    if request.method == 'POST':
        employee_id = request.POST.get('employee')
        date = request.POST.get('date')
        check_in = request.POST.get('check_in') or None
        check_out = request.POST.get('check_out') or None
        status = request.POST.get('status', 'present')
        note = request.POST.get('note', '')

        obj, created = Attendance.objects.update_or_create(
            employee_id=employee_id,
            date=date,
            defaults={'check_in': check_in, 'check_out': check_out, 'status': status, 'note': note}
        )
        messages.success(request, 'Attendance marked successfully.')
        return redirect('attendance_list')

    employees = Employee.objects.filter(status='active')
    return render(request, 'attendance/mark.html', {'employees': employees, 'today': datetime.date.today()})


@login_required
def attendance_delete(request, pk):
    attendance = get_object_or_404(Attendance, pk=pk)
    if request.method == 'POST':
        attendance.delete()
        messages.success(request, 'Attendance record deleted.')
        return redirect('attendance_list')
    return render(request, 'attendance/confirm_delete.html', {'attendance': attendance})


@login_required
def attendance_export(request):
    date_filter = request.GET.get('date', str(datetime.date.today()))
    attendances = Attendance.objects.filter(date=date_filter).select_related('employee')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance {date_filter}"

    header_fill = PatternFill("solid", fgColor="059669")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center      = Alignment(horizontal="center", vertical="center")
    border      = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'),  bottom=Side(style='thin'))

    headers = ['#','Employee ID','Name','Date','Check In','Check Out','Status','Note']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = center; cell.border = border

    alt_fill = PatternFill("solid", fgColor="f0fdf4")
    for i, a in enumerate(attendances, 1):
        ws.append([
            i, a.employee.employee_id, a.employee.name,
            str(a.date), str(a.check_in or '—'), str(a.check_out or '—'),
            a.get_status_display(), a.note or '—'
        ])
        for cell in ws[i+1]:
            cell.border = border; cell.alignment = Alignment(vertical="center")
            if i % 2 == 0: cell.fill = alt_fill

    for col, w in zip(['A','B','C','D','E','F','G','H'],
                      [5, 14, 22, 12, 10, 10, 12, 20]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Attendance_{date_filter}.xlsx"'
    wb.save(response)
    return response
